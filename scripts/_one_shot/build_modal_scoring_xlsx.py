#!/usr/bin/env python3
"""Build a visual Excel workbook for manual modal LoRA scoring."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_PATH = Path("docs/Modal_LoRA_Manual_Scoring_Visual.xlsx")

RUN_GROUPS = [
    ("r128_a256_d005", "escucha", ["epoch_20_loss_0.1982", "epoch_100_loss_0.1788", "epoch_140_loss_0.1714"]),
    ("r128_a256_d005", "reserva", ["epoch_180_loss_0.1603"]),
    ("r64_a128_d0", "escucha", ["epoch_160_loss_0.1725", "epoch_180_loss_0.1748", "epoch_80_loss_0.1921"]),
    ("r64_a128_d0", "reserva", ["epoch_200_loss_0.1704"]),
    ("r64_a128_d005", "escucha", ["epoch_160_loss_0.1766", "epoch_100_loss_0.1843", "epoch_200_loss_0.1701"]),
    ("r64_a128_d005", "reserva", ["epoch_120_loss_0.1808"]),
    ("r64_a128_d005_lr5e-5", "escucha", ["epoch_100_loss_0.1965", "epoch_160_loss_0.1952", "epoch_140_loss_0.1960"]),
    ("r64_a128_d005_lr5e-5", "reserva", ["epoch_200_loss_0.1936"]),
]

MODES = ["ionian", "dorian", "phrygian", "lydian", "mixolydian", "aeolian", "locrian"]
SEEDS = [1111, 2222]

HEADERS = [
    "Run",
    "Tipo",
    "Checkpoint",
    "Modo",
    "Seed",
    "Centro_tonal_1_5",
    "Color_modal_1_5",
    "No_modulacion_1_5",
    "Coherencia_1_5",
    "Contraste_1_5",
    "Total_0_25",
    "Prom_ckpt_total",
    "Prom_ckpt_modo_total",
    "Decision_ckpt",
    "Notas",
]


def _build_rows() -> list[list[str | int]]:
    rows: list[list[str | int]] = []
    row_idx = 2
    for run, tipo, checkpoints in RUN_GROUPS:
        for ckpt in checkpoints:
            for mode in MODES:
                for seed in SEEDS:
                    rows.append(
                        [
                            run,
                            tipo,
                            ckpt,
                            mode,
                            seed,
                            "",
                            "",
                            "",
                            "",
                            "",
                            f'=IF(COUNTA(F{row_idx}:J{row_idx})=0,"",SUM(F{row_idx}:J{row_idx}))',
                            f'=IF(K{row_idx}="","",AVERAGEIFS($K:$K,$A:$A,$A{row_idx},$C:$C,$C{row_idx}))',
                            f'=IF(K{row_idx}="","",AVERAGEIFS($K:$K,$A:$A,$A{row_idx},$C:$C,$C{row_idx},$D:$D,$D{row_idx}))',
                            "",
                            "",
                        ]
                    )
                    row_idx += 1
    return rows


def _style_scoring_sheet(ws, last_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{last_row}"

    widths = {
        "A": 22,
        "B": 12,
        "C": 24,
        "D": 14,
        "E": 8,
        "F": 18,
        "G": 18,
        "H": 20,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 16,
        "M": 19,
        "N": 18,
        "O": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=15):
        for cell in row:
            if cell.column in [1, 2, 3, 4, 5, 14, 15]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    score_validation = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
    score_validation.error = "Introduce un entero entre 1 y 5"
    score_validation.errorTitle = "Valor no valido"
    ws.add_data_validation(score_validation)
    score_validation.add(f"F2:J{last_row}")

    decision_validation = DataValidation(
        type="list",
        formula1='"descartar,candidato_debil,candidato_fuerte,finalista"',
    )
    decision_validation.error = "Selecciona una opcion de la lista"
    decision_validation.errorTitle = "Decision no valida"
    ws.add_data_validation(decision_validation)
    decision_validation.add(f"N2:N{last_row}")

    score_colors = ColorScaleRule(
        start_type="num",
        start_value=1,
        start_color="F8696B",
        mid_type="num",
        mid_value=3,
        mid_color="FFEB84",
        end_type="num",
        end_value=5,
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"F2:J{last_row}", score_colors)

    total_colors = ColorScaleRule(
        start_type="num",
        start_value=5,
        start_color="F8696B",
        mid_type="num",
        mid_value=15,
        mid_color="FFEB84",
        end_type="num",
        end_value=25,
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"K2:M{last_row}", total_colors)

    escucha_fill = PatternFill("solid", fgColor="E8F4FF")
    reserva_fill = PatternFill("solid", fgColor="FFF5E6")
    ws.conditional_formatting.add(
        f"A2:O{last_row}",
        FormulaRule(formula=["$B2=\"escucha\""], stopIfTrue=False, fill=escucha_fill),
    )
    ws.conditional_formatting.add(
        f"A2:O{last_row}",
        FormulaRule(formula=["$B2=\"reserva\""], stopIfTrue=False, fill=reserva_fill),
    )

    table = Table(displayName="ScoringTable", ref=f"A1:O{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _build_summary_sheet(wb, rows: list[list[str | int]]) -> None:
    ws = wb.create_sheet("Resumen_ckpt")
    headers = [
        "Run",
        "Tipo",
        "Checkpoint",
        "Promedio_checkpoint",
        "Decision",
        "Notas",
    ]
    ws.append(headers)

    seen: set[tuple[str, str, str]] = set()
    unique: list[tuple[str, str, str]] = []
    for row in rows:
        key = (str(row[0]), str(row[1]), str(row[2]))
        if key not in seen:
            seen.add(key)
            unique.append(key)

    for idx, (run, tipo, ckpt) in enumerate(unique, start=2):
        ws.append(
            [
                run,
                tipo,
                ckpt,
                f'=AVERAGEIFS(Scoring!$K:$K,Scoring!$A:$A,A{idx},Scoring!$C:$C,C{idx})',
                "",
                "",
            ]
        )

    header_fill = PatternFill("solid", fgColor="375623")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 24, "B": 12, "C": 24, "D": 20, "E": 18, "F": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    last_row = ws.max_row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{last_row}"

    color_rule = ColorScaleRule(
        start_type="num",
        start_value=5,
        start_color="F8696B",
        mid_type="num",
        mid_value=15,
        mid_color="FFEB84",
        end_type="num",
        end_value=25,
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"D2:D{last_row}", color_rule)

    decision_validation = DataValidation(
        type="list",
        formula1='"descartar,candidato_debil,candidato_fuerte,finalista"',
    )
    ws.add_data_validation(decision_validation)
    decision_validation.add(f"E2:E{last_row}")


def _build_legend_sheet(wb) -> None:
    ws = wb.create_sheet("Leyenda")
    ws.append(["Campo", "Significado breve"])
    items = [
        ("Centro_tonal_1_5", "Claridad y estabilidad de la tonica."),
        ("Color_modal_1_5", "Reconocimiento del modo objetivo."),
        ("No_modulacion_1_5", "Ausencia de cambios tonales no deseados."),
        ("Coherencia_1_5", "Continuidad musical global."),
        ("Contraste_1_5", "Diferencia perceptible frente a otros modos."),
        ("Total_0_25", "Suma de los cinco criterios (maximo 25)."),
    ]
    for item in items:
        ws.append(item)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 64
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="7F6000")
    ws["B1"].fill = PatternFill("solid", fgColor="7F6000")


def main() -> int:
    rows = _build_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scoring"

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    last_row = ws.max_row
    _style_scoring_sheet(ws, last_row)
    _build_summary_sheet(wb, rows)
    _build_legend_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH} | rows={len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
